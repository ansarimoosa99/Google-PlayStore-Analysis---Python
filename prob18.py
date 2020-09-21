import pandas as pd
from openpyxl import load_workbook
from textblob import TextBlob

def add_data(strr):
    
    data=pd.read_excel("App-data.xlsx")

    filepath='App-data.xlsx'
    
    wb=load_workbook(filepath)
    
    sheet=wb.active
    
    #enter=['PUBG','GAME','NAN']
    
    maxx=(len(data['App']))+1
    #print(maxx)
    
    c=1
    for item in range(len(strr)):
        sheet.cell(row=maxx+1, column=c).value=strr[item]
        c=c+1 
     
    wb.save(filepath)
    
    lst=[]
    lst.append("The data is sucessfully added")
    return lst

def user_review(appnreview):
    strr=[]
    for k,v in appnreview.items():
        analysis=TextBlob(v)
        print(analysis.sentiment)
        strr.append(k)
        strr.append(v)
        if analysis.sentiment.polarity<0:
            strr.append("Negative")
        elif analysis.sentiment.polarity==0:
            strr.append("Neutral")
        elif analysis.sentiment.polarity>0:
            strr.append("Positive")
        else:
             strr.append("nan")
        strr.append(analysis.sentiment.polarity)
        strr.append(analysis.sentiment.subjectivity)
    print(strr)
    data=pd.read_excel("user_reviews.xlsx")

    filepath='user_reviews.xlsx'
    wb=load_workbook(filepath)
    
    sheet=wb.active
    
    #enter=['PUBG','GAME','NAN']
    
    maxx=(len(data['App']))+1
    #print(maxx)
    
    c=1
    for item in range(len(strr)):
        sheet.cell(row=maxx+1, column=c).value=strr[item]
        c=c+1 
     
    wb.save(filepath)
    
    lst=[]
    lst.append("The data is sucessfully added")
    return lst

    